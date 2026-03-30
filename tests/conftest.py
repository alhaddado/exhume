import pytest
import struct
from pathlib import Path
from openpyxl import Workbook as OpenpyxlWorkbook


@pytest.fixture
def tmp_xlsx(tmp_path) -> Path:
    """Create a minimal .xlsx with cell data for testing."""
    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Value"
    ws["A2"] = 1
    ws["B2"] = "Alpha"
    ws["C2"] = 10.5
    ws["A3"] = 2
    ws["B3"] = "Beta"
    ws["C3"] = 20.0
    ws["D3"] = "=C2+C3"

    ws.merge_cells("A5:C5")
    ws["A5"] = "Merged Header"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Second"

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def ole10native_data() -> tuple[str, bytes, bytes]:
    """Create a minimal Ole10Native stream payload.

    Returns (filename, raw_content, ole10native_bytes).
    The ole10native_bytes is the content of the Ole10Native stream
    inside a Packager Shell Object .bin file.
    """
    filename = "sample.txt"
    src_path = "C:\\Users\\test\\sample.txt"
    temp_path = "C:\\Temp\\sample.txt"
    content = b"https://api.example.com/test\r\n\r\n{\"key\": \"value\"}"

    # Build Ole10Native stream:
    # 2 bytes flags + filename\0 + src_path\0 + 4 bytes reserved + temp_path\0 + 4 bytes data_size + data
    body = b""
    body += struct.pack("<H", 2)  # flags
    body += filename.encode("ascii") + b"\x00"
    body += src_path.encode("ascii") + b"\x00"
    body += struct.pack("<I", 0)  # reserved
    body += temp_path.encode("ascii") + b"\x00"
    body += struct.pack("<I", len(content))
    body += content

    # Prepend total size (4 bytes)
    ole10native_bytes = struct.pack("<I", len(body)) + body

    return filename, content, ole10native_bytes
