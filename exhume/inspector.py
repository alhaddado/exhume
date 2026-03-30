"""Sheet and cell introspection via openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from exhume.models import CellInfo, SheetInfo, WorkbookInfo

# openpyxl type codes to human-readable names
_TYPE_MAP = {"s": "string", "n": "number", "b": "boolean", "d": "date", "e": "error", "f": "formula"}


def _cell_type(cell: Cell) -> str:
    if cell.data_type == "f":
        # For formula cells, infer the value type
        return "number" if isinstance(cell.value, (int, float)) else "string"
    return _TYPE_MAP.get(cell.data_type, "string")


def _cell_info(cell: Cell) -> CellInfo:
    formula = None
    value = cell.value
    if cell.data_type == "f":
        formula = f"={cell.value}" if not str(cell.value).startswith("=") else str(cell.value)
        value = cell.value

    return CellInfo(
        coordinate=cell.coordinate,
        value=value,
        data_type=_cell_type(cell),
        formula=formula,
        number_format=cell.number_format,
    )


def _resolve_sheet(wb, sheet_ref: Union[str, int]) -> Worksheet:
    if isinstance(sheet_ref, int):
        if sheet_ref < 0 or sheet_ref >= len(wb.sheetnames):
            raise ValueError(f"Sheet index {sheet_ref} not found (workbook has {len(wb.sheetnames)} sheets)")
        return wb[wb.sheetnames[sheet_ref]]
    if sheet_ref in wb.sheetnames:
        return wb[sheet_ref]
    raise ValueError(f"Sheet '{sheet_ref}' not found. Available: {wb.sheetnames}")


def inspect_workbook(path: Union[str, Path]) -> WorkbookInfo:
    """Return structural information about the workbook."""
    path = Path(path)
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = []
        for idx, name in enumerate(wb.sheetnames):
            ws = wb[name]
            merged = [str(m) for m in ws.merged_cells.ranges] if hasattr(ws, "merged_cells") else []
            sheets.append(SheetInfo(
                name=name,
                index=idx,
                dimensions=ws.dimensions or "",
                max_row=ws.max_row or 0,
                max_column=ws.max_column or 0,
                merged_cells=merged,
            ))
        return WorkbookInfo(
            filename=path.name,
            sheet_count=len(wb.sheetnames),
            sheets=sheets,
            named_ranges=list(wb.defined_names.keys()) if hasattr(wb, "defined_names") else [],
        )
    finally:
        wb.close()


def inspect_sheet_cells(
    path: Union[str, Path],
    sheet_ref: Union[str, int],
) -> list[CellInfo]:
    """Return all non-empty cells in a sheet."""
    path = Path(path)
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ws = _resolve_sheet(wb, sheet_ref)
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells.append(_cell_info(cell))
        return cells
    finally:
        wb.close()
