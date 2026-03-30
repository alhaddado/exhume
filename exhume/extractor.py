"""OLE object extraction from .xlsx files.

Pipeline:
1. Unzip .xlsx -> parse xl/drawings/vmlDrawing*.vml for shape anchors
2. Parse xl/drawings/_rels/ for shape-to-image mappings
3. Parse xl/worksheets/_rels/ for shape-to-oleObject mappings
4. Parse xl/worksheets/sheet*.xml for <oleObject> entries (shapeId -> rId)
5. Read each .bin via olefile -> extract Ole10Native stream
6. Strip Packager Shell Object header -> clean content
"""

from __future__ import annotations

import re
import struct
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Union
from zipfile import ZipFile

import olefile
from openpyxl import load_workbook

from exhume.models import EmbeddedObject

NS_V = "urn:schemas-microsoft-com:vml"
NS_O = "urn:schemas-microsoft-com:office:office"
NS_X = "urn:schemas-microsoft-com:office:excel"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def parse_ole10native(data: bytes, clean: bool = False) -> tuple[str, bytes]:
    """Parse an Ole10Native stream and return (filename, content).

    Args:
        data: Raw bytes of the Ole10Native stream.
        clean: If True, strip any binary path prefix before the URL in content.
    """
    pos = 4  # skip total_size (4 bytes)
    pos += 2  # skip flags (2 bytes)

    # Read null-terminated filename
    end = data.index(b"\x00", pos)
    filename = data[pos:end].decode("ascii", errors="replace")
    pos = end + 1

    # Read null-terminated source path
    end = data.index(b"\x00", pos)
    pos = end + 1

    # Skip reserved 4 bytes
    pos += 4

    # Read null-terminated temp path
    end = data.index(b"\x00", pos)
    pos = end + 1

    # Read data size and content
    data_size = struct.unpack_from("<I", data, pos)[0]
    pos += 4
    content = data[pos : pos + data_size]

    if clean:
        # Strip binary prefix before the URL
        url_match = re.search(rb"https?://\S+", content)
        if url_match:
            content = content[url_match.start() :]

    return filename, content


def extract_object_content(data: bytes) -> dict[str, Any]:
    """Parse Ole10Native stream bytes and return filename + raw content."""
    filename, content = parse_ole10native(data)
    return {"filename": filename, "content": content}


def _parse_vml_shapes(zf: ZipFile) -> dict[int, dict[str, Any]]:
    """Parse VML drawings to get shape anchors (row/col positions).

    Returns dict keyed by shape_id numeric value -> {row, col, draw_rid}.
    """
    shapes: dict[int, dict[str, Any]] = {}

    vml_files = [n for n in zf.namelist() if "vmlDrawing" in n and n.endswith(".vml")]
    for vml_path in vml_files:
        root = ET.fromstring(zf.read(vml_path))
        for shape in root.iter(f"{{{NS_V}}}shape"):
            shape_id_str = shape.get("id", "")
            if not shape_id_str.startswith("_x0000_s"):
                continue
            shape_id = int(shape_id_str.replace("_x0000_s", ""))

            imagedata = shape.find(f"{{{NS_V}}}imagedata")
            draw_rid = imagedata.get(f"{{{NS_O}}}relid", "") if imagedata is not None else ""

            client_data = shape.find(f"{{{NS_X}}}ClientData")
            if client_data is None:
                continue
            anchor_elem = client_data.find(f"{{{NS_X}}}Anchor")
            if anchor_elem is None or not anchor_elem.text:
                continue

            parts = [int(x.strip()) for x in anchor_elem.text.strip().split(",")]
            # Anchor format: col1, col1off, row1, row1off, col2, col2off, row2, row2off
            shapes[shape_id] = {
                "row": parts[2],  # 0-indexed
                "col": parts[0],
                "draw_rid": draw_rid,
            }

    return shapes


def _parse_worksheet_ole_mappings(zf: ZipFile) -> dict[int, str]:
    """Parse worksheet XML to map shapeId -> oleObject .bin filename.

    Returns dict keyed by shape_id -> relative path like '../embeddings/oleObject1.bin'.
    """
    # First, collect rId -> target from worksheet relationship files
    ws_rels_files = [n for n in zf.namelist() if "worksheets/_rels/" in n and n.endswith(".rels")]

    rid_to_target: dict[str, str] = {}
    for rels_path in ws_rels_files:
        root = ET.fromstring(zf.read(rels_path))
        for rel in root.iter(f"{{{NS_PKG_REL}}}Relationship"):
            if "oleObject" in rel.get("Type", ""):
                rid_to_target[rel.get("Id", "")] = rel.get("Target", "")

    # Then, parse sheet XML for <oleObject> entries
    shape_to_ole: dict[int, str] = {}
    sheet_files = [n for n in zf.namelist() if "worksheets/sheet" in n and n.endswith(".xml")]
    seen: set[tuple[int, str]] = set()

    for sheet_path in sheet_files:
        root = ET.fromstring(zf.read(sheet_path))
        for obj in root.iter(f"{{{NS_MAIN}}}oleObject"):
            shape_id = int(obj.get("shapeId", "0"))
            rid = obj.get(f"{{{NS_R}}}id", "")
            key = (shape_id, rid)
            if key not in seen and rid in rid_to_target:
                seen.add(key)
                shape_to_ole[shape_id] = rid_to_target[rid]

    return shape_to_ole


def _get_neighbor_cells(wb, sheet_name: str, row: int, max_cols: int = 10) -> dict[str, Any]:
    """Read cell values from neighboring columns in the same row."""
    ws = wb[sheet_name]
    neighbors: dict[str, Any] = {}
    for col_idx in range(1, max_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            neighbors[cell.coordinate] = cell.value
    return neighbors


def extract_objects_metadata(
    path: Union[str, Path],
    sheet_name: str | None = None,
) -> list[EmbeddedObject]:
    """List all embedded OLE objects in the workbook with metadata.

    Args:
        path: Path to the .xlsx file.
        sheet_name: Optional sheet filter.
    """
    path = Path(path)

    with ZipFile(path, "r") as zf:
        # Check if there are any OLE objects at all
        has_ole = any("oleObject" in n for n in zf.namelist())
        if not has_ole:
            return []

        shapes = _parse_vml_shapes(zf)
        shape_to_ole = _parse_worksheet_ole_mappings(zf)

    # Read cell values for neighbor context
    wb = load_workbook(path, read_only=False, data_only=True)
    sheet_names = wb.sheetnames

    objects: list[EmbeddedObject] = []
    idx = 0

    # Sort by shape_id for consistent ordering
    for shape_id in sorted(shapes.keys()):
        if shape_id not in shape_to_ole:
            continue

        shape = shapes[shape_id]
        ole_target = shape_to_ole[shape_id]
        ole_filename = ole_target.split("/")[-1]

        # Determine which sheet this object belongs to
        # For now, use the first sheet (most common case)
        obj_sheet = sheet_names[0] if sheet_names else "Sheet1"

        if sheet_name and obj_sheet != sheet_name:
            continue

        # Row in VML is 0-indexed, convert to 1-indexed for openpyxl
        row_1indexed = shape.get("row", 0) + 1

        # Get neighbor cell values
        neighbors = _get_neighbor_cells(wb, obj_sheet, row_1indexed)

        # Try to extract filename from the OLE object
        embedded_filename = ole_filename
        ole_path = f"xl/embeddings/{ole_filename}"
        try:
            with ZipFile(path, "r") as zf:
                if ole_path in zf.namelist():
                    ole_data = zf.read(ole_path)
                    ole_file = olefile.OleFileIO(ole_data)
                    if ole_file.exists("\x01Ole10Native"):
                        stream_data = ole_file.openstream("\x01Ole10Native").read()
                        embedded_filename, content = parse_ole10native(stream_data)
                        size = len(content)
                    else:
                        size = len(ole_data)
                    ole_file.close()
                else:
                    size = 0
        except Exception:
            size = 0

        idx += 1
        objects.append(EmbeddedObject(
            index=idx,
            filename=embedded_filename,
            sheet=obj_sheet,
            row=row_1indexed,
            column=shape.get("col", 0),
            shape_id=shape_id,
            ole_source=ole_filename,
            size_bytes=size,
            prog_id="Packager Shell Object",
            neighbor_cells=neighbors,
        ))

    wb.close()
    return objects


def extract_objects_to_disk(
    path: Union[str, Path],
    out_dir: Union[str, Path],
    by_row: bool = False,
    clean: bool = True,
) -> list[EmbeddedObject]:
    """Extract all embedded OLE objects to disk.

    Args:
        path: Path to the .xlsx file.
        out_dir: Directory to write extracted files.
        by_row: If True, organize files as <out>/row-<N>/<filename>.
        clean: If True, strip binary path prefix from content.

    Returns list of EmbeddedObject with metadata for each extracted file.
    """
    path = Path(path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    objects = extract_objects_metadata(path)
    if not objects:
        return []

    # Track filenames for deduplication
    used_names: dict[str, int] = {}

    with ZipFile(path, "r") as zf:
        for obj in objects:
            ole_path = f"xl/embeddings/{obj.ole_source}"
            if ole_path not in zf.namelist():
                continue

            ole_data = zf.read(ole_path)
            try:
                ole_file = olefile.OleFileIO(ole_data)
                if not ole_file.exists("\x01Ole10Native"):
                    ole_file.close()
                    continue
                stream_data = ole_file.openstream("\x01Ole10Native").read()
                ole_file.close()
            except Exception:
                continue

            _, content = parse_ole10native(stream_data, clean=clean)

            # Determine output path
            fname = obj.filename
            if by_row:
                target_dir = out_dir / f"row-{obj.row}"
            else:
                target_dir = out_dir

            target_dir.mkdir(parents=True, exist_ok=True)

            # Deduplicate filenames
            key = str(target_dir / fname)
            if key in used_names:
                used_names[key] += 1
                stem = Path(fname).stem
                suffix = Path(fname).suffix
                fname = f"{stem}_{used_names[key]}{suffix}"
            else:
                used_names[key] = 1

            out_path = target_dir / fname
            out_path.write_bytes(content)

    return objects
