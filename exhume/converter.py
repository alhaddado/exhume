"""Full workbook conversion to CSV/JSON directory structure."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from exhume.inspector import inspect_workbook, inspect_sheet_cells
from exhume.extractor import extract_objects_metadata, extract_objects_to_disk
from exhume.output import format_json


def convert_workbook(
    path: Union[str, Path],
    out_dir: Union[str, Path],
    fmt: str = "json",
) -> None:
    """Convert workbook to a structured output directory.

    Args:
        path: Path to the .xlsx file.
        out_dir: Root output directory.
        fmt: "json" or "csv" for cell data format.
    """
    path = Path(path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    info = inspect_workbook(path)

    # Write workbook-level metadata
    meta_path = out_dir / "metadata.json"
    meta_path.write_text(format_json(info, pretty=True))

    # Write per-sheet data and metadata
    wb = load_workbook(path, read_only=False, data_only=False)
    for sheet_info in info.sheets:
        sheet_dir = out_dir / "sheets" / sheet_info.name
        sheet_dir.mkdir(parents=True, exist_ok=True)

        # Sheet metadata
        sheet_meta_path = sheet_dir / "metadata.json"
        sheet_meta_path.write_text(format_json(sheet_info, pretty=True))

        # Sheet data
        ws = wb[sheet_info.name]
        if fmt == "json":
            _write_sheet_json(ws, sheet_dir / "data.json")
        else:
            _write_sheet_csv(ws, sheet_dir / "data.csv")

    wb.close()

    # Extract embedded objects
    objects_dir = out_dir / "objects"
    objects_dir.mkdir(parents=True, exist_ok=True)
    files_dir = objects_dir / "files"

    objects = extract_objects_metadata(path)
    if objects:
        extract_objects_to_disk(path, files_dir, by_row=False, clean=True)

    manifest = {
        "totalObjects": len(objects),
        "objects": [obj.to_dict() for obj in objects],
    }
    (objects_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2, default=str, ensure_ascii=False)
    )


def _write_sheet_json(ws, out_path: Path) -> None:
    """Write sheet cell data as JSON."""
    max_col = ws.max_column or 1
    headers = [get_column_letter(i) for i in range(1, max_col + 1)]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = {}
        has_value = False
        for cell in row:
            if cell.value is not None:
                has_value = True
                formula = None
                if cell.data_type == "f":
                    formula = str(cell.value) if str(cell.value).startswith("=") else f"={cell.value}"

                cell_type = "string"
                if isinstance(cell.value, (int, float)):
                    cell_type = "number"
                elif isinstance(cell.value, bool):
                    cell_type = "boolean"

                cells[get_column_letter(cell.column)] = {
                    "value": cell.value,
                    "type": cell_type,
                    "formula": formula,
                }
        if has_value:
            rows.append({"rowIndex": row[0].row, "cells": cells})

    data = {"headers": headers, "rows": rows}
    out_path.write_text(json.dumps(data, indent=2, default=str, ensure_ascii=False))


def _write_sheet_csv(ws, out_path: Path) -> None:
    """Write sheet cell data as CSV (values only)."""
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
